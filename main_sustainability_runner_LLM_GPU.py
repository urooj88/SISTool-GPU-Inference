#!/usr/bin/env python3

import os
import sys
import subprocess
from copy import copy

import pandas as pd
from rich.console import Console
from rich.table import Table
from rich import box
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ----------------------------
# Paths
# ----------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEST_MODELS_DIR = os.path.join(BASE_DIR, "test_models")
BUILD_DATASET_SCRIPT = os.path.join(BASE_DIR, "build_eval_dataset.py")
OMEGAWATT_SCRIPT = os.path.join(BASE_DIR, "run_omegawatt_log_both_models12_same.sh")
COLLECT_SCRIPT = os.path.join(TEST_MODELS_DIR, "collect_inference_metrics.py")
INFERENCE_CSV = os.path.join(TEST_MODELS_DIR, "model_inference_data.csv")
METRICS_CSV = os.path.join(BASE_DIR, "model_metrices.csv")
MODEL_LOGS_DIR = os.path.join(BASE_DIR, "model_logs")

PROMPTS_FILE = os.path.join(TEST_MODELS_DIR, "prompts.txt")
ANSWERS_FILE = os.path.join(TEST_MODELS_DIR, "answers.csv")

# ----------------------------
# GPU setup fallback
# ----------------------------
NUM_GPUS = 1

# ----------------------------
# Setup
# ----------------------------
console = Console()
os.makedirs(MODEL_LOGS_DIR, exist_ok=True)


# ----------------------------
# Helpers
# ----------------------------
def normalize_series(series: pd.Series, higher_is_better: bool = True) -> pd.Series:
    series = pd.to_numeric(series, errors="coerce")
    min_val = series.min()
    max_val = series.max()

    if pd.isna(min_val) or pd.isna(max_val) or max_val == min_val:
        return pd.Series([0.5] * len(series), index=series.index)

    norm = (series - min_val) / (max_val - min_val)
    return norm if higher_is_better else 1 - norm


def classify_sis(score: float) -> str:
    if score <= 0.3:
        return "🟢 Low"
    elif score <= 0.7:
        return "🟡 Medium"
    else:
        return "🔴 High"


def dataset_exists() -> bool:
    return os.path.exists(PROMPTS_FILE) and os.path.exists(ANSWERS_FILE)


# ----------------------------
# Main
# ----------------------------
def main():
    # ----------------------------
    # STEP 0: Build dataset only if missing
    # ----------------------------
    console.print("[bold cyan]Checking evaluation dataset...[/bold cyan]")

    if not os.path.exists(BUILD_DATASET_SCRIPT):
        raise FileNotFoundError(f"Missing script: {BUILD_DATASET_SCRIPT}")

    if dataset_exists():
        console.print("[bold green]Using existing dataset files.[/bold green]")
        console.print(f"Prompts: {PROMPTS_FILE}")
        console.print(f"Answers: {ANSWERS_FILE}")
    else:
        console.print("[bold yellow]Dataset files not found. Building dataset...[/bold yellow]")
        subprocess.run([sys.executable, BUILD_DATASET_SCRIPT], cwd=BASE_DIR, check=True)
        console.print("[bold green]Dataset generation completed.[/bold green]")

    # ----------------------------
    # STEP 1: Run models + power logging
    # ----------------------------
    console.print("[bold cyan]Running model inference with power logging...[/bold cyan]")

    if not os.path.exists(OMEGAWATT_SCRIPT):
        raise FileNotFoundError(f"Missing script: {OMEGAWATT_SCRIPT}")

    subprocess.run(["bash", OMEGAWATT_SCRIPT], cwd=BASE_DIR, check=True)

    console.print("[bold green]Model execution completed.[/bold green]")

    # ----------------------------
    # STEP 2: Collect inference metrics
    # ----------------------------
    console.print("[bold cyan]Collecting inference metrics...[/bold cyan]")

    if not os.path.exists(COLLECT_SCRIPT):
        raise FileNotFoundError(f"Missing script: {COLLECT_SCRIPT}")

    subprocess.run([sys.executable, "collect_inference_metrics.py"], cwd=TEST_MODELS_DIR, check=True)

    console.print("[bold green]Inference metrics collection completed.[/bold green]")

    # ----------------------------
    # Load CSV files
    # ----------------------------
    if not os.path.exists(METRICS_CSV):
        raise FileNotFoundError(f"Missing metrics CSV: {METRICS_CSV}")

    if not os.path.exists(INFERENCE_CSV):
        raise FileNotFoundError(f"Missing inference CSV: {INFERENCE_CSV}")

    metrics_df = pd.read_csv(METRICS_CSV)
    inference_df = pd.read_csv(INFERENCE_CSV)

    # ----------------------------
    # Normalize model names
    # ----------------------------
    inference_df = inference_df.copy()
    if len(inference_df) >= 1:
        inference_df.loc[inference_df.index[0], "Model_Name"] = "Model1"
    if len(inference_df) >= 2:
        inference_df.loc[inference_df.index[1], "Model_Name"] = "Model2"
    if len(inference_df) >= 3:
        inference_df.loc[inference_df.index[2], "Model_Name"] = "Model3"
    if len(inference_df) >= 4:
        inference_df.loc[inference_df.index[3], "Model_Name"] = "Model4"

    # ----------------------------
    # Safe rename metrics columns
    # ----------------------------
    metrics_rename_map = {}
    if "Dynamic_Energy_J" in metrics_df.columns:
        metrics_rename_map["Dynamic_Energy_J"] = "Dynamic_Energy"
    if "Mem_Use (GiB)" in metrics_df.columns:
        metrics_rename_map["Mem_Use (GiB)"] = "Memory_GB"
    elif "Max_Mem_GiB" in metrics_df.columns:
        metrics_rename_map["Max_Mem_GiB"] = "Memory_GB"
    if "CPU_Core_Hours" in metrics_df.columns:
        metrics_rename_map["CPU_Core_Hours"] = "CPU_hours"
    if "GPU_Hours" in metrics_df.columns:
        metrics_rename_map["GPU_Hours"] = "GPU_hours"

    metrics_df.rename(columns=metrics_rename_map, inplace=True)

    # ----------------------------
    # Safe rename inference columns
    # ----------------------------
    inference_df.rename(
        columns={
            "Model_Acuracy (%)": "Model_Accuracy",
            "Model_Accuracy (%)": "Model_Accuracy",
            "Reasoning_Accuracy (%)": "Reasoning_Accuracy",
            "MCQ_Accuracy (%)": "MCQ_Accuracy",
            "Truth_Accuracy (%)": "Truth_Accuracy",
            "GSM8K_Accuracy (%)": "GSM8K_Accuracy",
            "MMLU_Accuracy (%)": "MMLU_Accuracy",
            "TruthfulQA_Accuracy (%)": "TruthfulQA_Accuracy",
            "queries": "Queries",
            "Queries_Run": "Queries",
            "T_Model_Size (MB)": "Model_Size_MB",
            "Model_Size_(MB)": "Model_Size_MB",
            "Learned_Parameters": "Parameters",
            "Flops_per_Inference": "FLOPs",
            "FLOPs_per_Inference": "FLOPs",
            "Carbon _Intensity (gCO₂/kWh)": "Carbon_Intensity",
            "Carbon_Intensity_(gCO2/kWh)": "Carbon_Intensity",
            "Generated_Tokens": "Output_Tokens",
            "Total_Generated_Tokens": "Output_Tokens",
            "Total_Output_Tokens": "Output_Tokens",
            "Completion_Tokens": "Output_Tokens",
            "Generated_Output_Tokens": "Output_Tokens",
        },
        inplace=True,
    )

    # ----------------------------
    # Merge
    # ----------------------------
    merged_df = pd.merge(metrics_df, inference_df, on="Model_Name", how="left")

    # ----------------------------
    # Required columns
    # ----------------------------
    required_cols = [
        "Dynamic_Energy",
        "Total_Execution_Time_s",
        "Queries",
        "Model_Accuracy",
        "Carbon_Intensity",
        "Output_Tokens",
    ]

    missing = [c for c in required_cols if c not in merged_df.columns]
    if missing:
        raise KeyError(f"Missing required columns after merge: {missing}")

    # ----------------------------
    # Numeric conversion
    # ----------------------------
    numeric_cols = [
        "Dynamic_Energy",
        "Total_Execution_Time_s",
        "Queries",
        "Model_Accuracy",
        "Carbon_Intensity",
        "Output_Tokens",
        "Reasoning_Accuracy",
        "MCQ_Accuracy",
        "Truth_Accuracy",
        "GSM8K_Accuracy",
        "MMLU_Accuracy",
        "TruthfulQA_Accuracy",
    ]

    for col in numeric_cols:
        if col not in merged_df.columns:
            merged_df[col] = 0.0
        merged_df[col] = pd.to_numeric(merged_df[col], errors="coerce")

    # ----------------------------
    # Safety checks
    # ----------------------------
    merged_df["Queries"] = merged_df["Queries"].fillna(0).astype(float)
    if (merged_df["Queries"] <= 0).any():
        raise ValueError("Queries must be > 0 for all models.")

    if (merged_df["Total_Execution_Time_s"] <= 0).any():
        raise ValueError("Total_Execution_Time_s must be > 0 for all models.")

    if (merged_df["Dynamic_Energy"] <= 0).any():
        raise ValueError("Dynamic_Energy must be > 0 for all models.")

    # ----------------------------
    # Optional / fallback fields
    # ----------------------------
    if "Memory_GB" not in merged_df.columns:
        merged_df["Memory_GB"] = 0.0
    merged_df["Memory_GB"] = pd.to_numeric(
        merged_df["Memory_GB"], errors="coerce"
    ).fillna(0.0)

    for col in [
        "CPU_hours",
        "GPU_Count",
        "GPU_hours",
        "FLOPs",
        "Model_Size_MB",
        "Parameters",
        "Avg_Output_Tokens_per_Query",
    ]:
        if col not in merged_df.columns:
            merged_df[col] = 0.0
        merged_df[col] = pd.to_numeric(merged_df[col], errors="coerce").fillna(0.0)

    merged_df["Reasoning_Accuracy"] = merged_df["Reasoning_Accuracy"].fillna(0.0)
    merged_df["MCQ_Accuracy"] = merged_df["MCQ_Accuracy"].fillna(0.0)
    merged_df["Truth_Accuracy"] = merged_df["Truth_Accuracy"].fillna(0.0)
    merged_df["GSM8K_Accuracy"] = merged_df["GSM8K_Accuracy"].fillna(0.0)
    merged_df["MMLU_Accuracy"] = merged_df["MMLU_Accuracy"].fillna(0.0)
    merged_df["TruthfulQA_Accuracy"] = merged_df["TruthfulQA_Accuracy"].fillna(0.0)

    # ----------------------------
    # GPU hours
    # ----------------------------
    # Prefer GPU_hours already computed by the OmegaWatt script.
    # If missing or zero, compute from GPU_Count.
    # If GPU_Count is missing too, use NUM_GPUS fallback.
    computed_gpu_hours = merged_df["GPU_Count"] * merged_df["Total_Execution_Time_s"] / 3600.0
    fallback_gpu_hours = NUM_GPUS * merged_df["Total_Execution_Time_s"] / 3600.0

    merged_df["GPU_hours"] = merged_df["GPU_hours"].where(merged_df["GPU_hours"] > 0, computed_gpu_hours)
    merged_df["GPU_hours"] = merged_df["GPU_hours"].where(merged_df["GPU_hours"] > 0, fallback_gpu_hours)
    merged_df["GPU_Count"] = merged_df["GPU_Count"].where(merged_df["GPU_Count"] > 0, NUM_GPUS)

    if (merged_df["GPU_hours"] <= 0).any():
        raise ValueError("GPU_hours must be > 0 for all models.")

    # ----------------------------
    # Carbon emissions
    # ----------------------------
    merged_df["Carbon_Emissions"] = (
        merged_df["Dynamic_Energy"] * merged_df["Carbon_Intensity"] / 3_600_000
    )

    # ----------------------------
    # Sustainability metrics
    # ----------------------------
    merged_df["Runtime_per_Query"] = merged_df["Total_Execution_Time_s"] / merged_df["Queries"]
    merged_df["Energy_per_Query"] = merged_df["Dynamic_Energy"] / merged_df["Queries"]
    merged_df["CO2_per_Query"] = merged_df["Carbon_Emissions"] / merged_df["Queries"]

    merged_df["Acc_per_Energy"] = (
        merged_df["Model_Accuracy"] / merged_df["Dynamic_Energy"].replace(0, pd.NA)
    )
    merged_df["Hardware_Efficiency"] = (
        merged_df["Model_Accuracy"] / merged_df["GPU_hours"].replace(0, pd.NA)
    )

    merged_df["Tokens_per_Query"] = merged_df["Output_Tokens"] / merged_df["Queries"]
    merged_df["Inference_Throughput"] = (
        merged_df["Output_Tokens"] / merged_df["Total_Execution_Time_s"]
    )
    merged_df["Token_Energy_Efficiency"] = (
        merged_df["Output_Tokens"] / merged_df["Dynamic_Energy"]
    )

    merged_df["Acc_per_Energy"] = merged_df["Acc_per_Energy"].fillna(0.0)
    merged_df["Hardware_Efficiency"] = merged_df["Hardware_Efficiency"].fillna(0.0)
    merged_df["Tokens_per_Query"] = merged_df["Tokens_per_Query"].fillna(0.0)
    merged_df["Inference_Throughput"] = merged_df["Inference_Throughput"].fillna(0.0)
    merged_df["Token_Energy_Efficiency"] = merged_df["Token_Energy_Efficiency"].fillna(0.0)

    # ----------------------------
    # Normalize metrics
    # ----------------------------
    merged_df["Energy_norm"] = normalize_series(merged_df["Energy_per_Query"], higher_is_better=False)
    merged_df["CO2_norm"] = normalize_series(merged_df["CO2_per_Query"], higher_is_better=False)
    merged_df["Runtime_norm"] = normalize_series(merged_df["Runtime_per_Query"], higher_is_better=False)
    merged_df["Memory_norm"] = normalize_series(merged_df["Memory_GB"], higher_is_better=False)
    merged_df["FLOPs_norm"] = normalize_series(merged_df["FLOPs"], higher_is_better=False)
    merged_df["Model_Size_norm"] = normalize_series(merged_df["Model_Size_MB"], higher_is_better=False)
    merged_df["Accuracy_norm"] = normalize_series(merged_df["Model_Accuracy"], higher_is_better=True)
    merged_df["Throughput_norm"] = normalize_series(merged_df["Inference_Throughput"], higher_is_better=True)
    merged_df["TEE_norm"] = normalize_series(merged_df["Token_Energy_Efficiency"], higher_is_better=True)

    # ----------------------------
    # SIS score
    # ----------------------------
    sis_components = [
        "Energy_norm",
        "CO2_norm",
        "Runtime_norm",
        "Memory_norm",
        "FLOPs_norm",
        "Model_Size_norm",
        "Accuracy_norm",
        "Throughput_norm",
        "TEE_norm",
    ]

    weights = {k: 1 / len(sis_components) for k in sis_components}

    merged_df["Sustainability_Goodness"] = sum(
        weights[k] * merged_df[k] for k in sis_components
    )
    merged_df["SIS"] = 1 - merged_df["Sustainability_Goodness"]
    merged_df["SIS"] = merged_df["SIS"].clip(0, 1)
    merged_df["Sustainability_Level"] = merged_df["SIS"].apply(classify_sis)

    # ----------------------------
    # Terminal table 1
    # ----------------------------
    classification_table = Table(
        title=" SIS Classification Ranges",
        box=box.SQUARE,
        show_lines=True
    )
    classification_table.add_column("SIS Range", justify="center")
    classification_table.add_column("Sustainability Level", justify="center")
    classification_table.add_row("0.0 – 0.3", "🟢 Low")
    classification_table.add_row("0.3 – 0.7", "🟡 Medium")
    classification_table.add_row("0.7 – 1.0", "🔴 High")

    # ----------------------------
    # Terminal table 2
    # ----------------------------
    summary_table = Table(
        title=" Sustainability Summary for Inference Models",
        box=box.ROUNDED,
        show_lines=True
    )
    summary_table.add_column("Model", style="cyan")
    summary_table.add_column("Accuracy (%)", justify="right")
    summary_table.add_column("Reason (%)", justify="right")
    summary_table.add_column("MCQ (%)", justify="right")
    summary_table.add_column("Truth (%)", justify="right")
    summary_table.add_column("Tokens/sec", justify="right")
    summary_table.add_column("Tokens/J", justify="right")
    summary_table.add_column("SIS Score", justify="right", style="magenta")
    summary_table.add_column("Level", justify="center")

    for _, row in merged_df.iterrows():
        summary_table.add_row(
            str(row["Model_Name"]),
            f"{row['Model_Accuracy']:.2f}",
            f"{row['Reasoning_Accuracy']:.2f}",
            f"{row['MCQ_Accuracy']:.2f}",
            f"{row['Truth_Accuracy']:.2f}",
            f"{row['Inference_Throughput']:.4f}",
            f"{row['Token_Energy_Efficiency']:.6f}",
            f"{row['SIS']:.4f}",
            str(row["Sustainability_Level"]),
        )

    # ----------------------------
    # Terminal table 3
    # ----------------------------
    details_table = Table(
        title=" Detailed Sustainability Metrics for Inference Models",
        box=box.ROUNDED,
        show_lines=True
    )
    details_table.add_column("Model", style="cyan")
    details_table.add_column("Energy/Query (J)", justify="right", style="yellow")
    details_table.add_column("gCO₂ eq/Query", justify="right", style="red")
    details_table.add_column("Acc/Energy", justify="right", style="green")
    details_table.add_column("Runtime/Query (s)", justify="right", style="blue")
    details_table.add_column("Output Tokens", justify="right")
    details_table.add_column("Tokens/Query", justify="right")
    details_table.add_column("Throughput (tok/s)", justify="right")
    details_table.add_column("TEE (tok/J)", justify="right")
    details_table.add_column("Model Size (MB)", justify="right")
    details_table.add_column("FLOPs", justify="right")
    details_table.add_column("Max Mem (GiB)", justify="right")
    details_table.add_column("GPU Count", justify="right")
    details_table.add_column("GPU Hours", justify="right")
    details_table.add_column("HWE (%/GPU_hr)", justify="right")

    for _, row in merged_df.iterrows():
        flops_str = "N/A"
        if pd.notna(row["FLOPs"]) and float(row["FLOPs"]) > 0:
            flops_str = str(int(row["FLOPs"]))

        details_table.add_row(
            str(row["Model_Name"]),
            f"{row['Energy_per_Query']:.4f}",
            f"{row['CO2_per_Query']:.8f}",
            f"{row['Acc_per_Energy']:.6f}",
            f"{row['Runtime_per_Query']:.6f}",
            f"{row['Output_Tokens']:.0f}",
            f"{row['Tokens_per_Query']:.4f}",
            f"{row['Inference_Throughput']:.4f}",
            f"{row['Token_Energy_Efficiency']:.6f}",
            f"{row['Model_Size_MB']:.2f}",
            flops_str,
            f"{row['Memory_GB']:.2f}",
            f"{row['GPU_Count']:.0f}",
            f"{row['GPU_hours']:.6f}",
            f"{row['Hardware_Efficiency']:.6f}",
        )

    console.print(classification_table)
    console.print(summary_table)
    console.print(details_table)

    # ----------------------------
    # Save text report
    # ----------------------------
    report_path = os.path.join(MODEL_LOGS_DIR, "sustainability_detailed_report.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        for _, row in merged_df.iterrows():
            f.write(f"\n=== Detailed Metrics for {row['Model_Name']} ===\n")
            f.write(f"Accuracy: {row['Model_Accuracy']:.2f}%\n")
            f.write(f"Reasoning Accuracy: {row['Reasoning_Accuracy']:.2f}%\n")
            f.write(f"MCQ Accuracy: {row['MCQ_Accuracy']:.2f}%\n")
            f.write(f"Truth Accuracy: {row['Truth_Accuracy']:.2f}%\n")
            f.write(f"GSM8K Accuracy: {row['GSM8K_Accuracy']:.2f}%\n")
            f.write(f"MMLU Accuracy: {row['MMLU_Accuracy']:.2f}%\n")
            f.write(f"TruthfulQA Accuracy: {row['TruthfulQA_Accuracy']:.2f}%\n")
            f.write(f"Total Output Tokens: {row['Output_Tokens']:.0f}\n")
            f.write(f"Tokens per Query: {row['Tokens_per_Query']:.4f}\n")
            f.write(f"Inference Throughput (tokens/sec): {row['Inference_Throughput']:.4f}\n")
            f.write(f"Token Energy Efficiency (tokens/J): {row['Token_Energy_Efficiency']:.6f}\n")
            f.write(f"Energy per Query: {row['Energy_per_Query']:.4f} J\n")
            f.write(f"CO2 per Query: {row['CO2_per_Query']:.8f} g\n")
            f.write(f"Accuracy per Energy: {row['Acc_per_Energy']:.6f} %/J\n")
            f.write(f"Runtime per Query: {row['Runtime_per_Query']:.6f} s\n")
            f.write(f"Model Size: {row['Model_Size_MB']:.2f} MB\n")
            f.write(
                f"FLOPs per Inference: "
                f"{str(int(row['FLOPs'])) if pd.notna(row['FLOPs']) and row['FLOPs'] > 0 else 'N/A'}\n"
            )
            f.write(f"Memory Usage (GiB): {row['Memory_GB']:.2f}\n")
            f.write(f"GPU Count: {row['GPU_Count']:.0f}\n")
            f.write(f"GPU Hours: {row['GPU_hours']:.6f}\n")
            f.write(f"Hardware Efficiency (%/GPU_hr): {row['Hardware_Efficiency']:.6f}\n")
            f.write(f"Sustainability Goodness: {row['Sustainability_Goodness']:.4f}\n")
            f.write(f"SIS (Impact): {row['SIS']:.4f}\n")
            f.write(f"Sustainability Level: {row['Sustainability_Level']}\n")
            f.write("-" * 60 + "\n")

    # ----------------------------
    # Save Excel
    # ----------------------------
    excel_file = os.path.join(MODEL_LOGS_DIR, "updated_sustainability_metrics.xlsx")
    merged_df.to_excel(excel_file, index=False)

    wb = load_workbook(excel_file)
    ws = wb.active

    green_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    highlight_cols = [
        "Reasoning_Accuracy",
        "MCQ_Accuracy",
        "Truth_Accuracy",
        "GSM8K_Accuracy",
        "MMLU_Accuracy",
        "TruthfulQA_Accuracy",
        "Runtime_per_Query",
        "Energy_per_Query",
        "CO2_per_Query",
        "Acc_per_Energy",
        "GPU_Count",
        "GPU_hours",
        "Hardware_Efficiency",
        "Tokens_per_Query",
        "Inference_Throughput",
        "Token_Energy_Efficiency",
        "Energy_norm",
        "CO2_norm",
        "Accuracy_norm",
        "Runtime_norm",
        "Memory_norm",
        "FLOPs_norm",
        "Model_Size_norm",
        "Throughput_norm",
        "TEE_norm",
        "Sustainability_Goodness",
        "SIS",
        "Sustainability_Level",
    ]

    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        header = col[0].value
        for cell in col:
            if header in highlight_cols:
                cell.fill = green_fill
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(i)].width = max(12, max_length + 2)

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.alignment:
                align = copy(cell.alignment)
                align.wrap_text = True
                cell.alignment = align

    wb.save(excel_file)

    console.print("[bold green]All results saved in the 'model_logs/' folder.[/bold green]")


if __name__ == "__main__":
    main()
